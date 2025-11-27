import os
import shutil
import openpyxl
import streamlit as st
from config import EXCEL_PATH, EXCEL_SHEET_NAME, EXCEL_INSERT_ROW


def save_to_excel(fields):
    if not os.path.exists(EXCEL_PATH):
        st.error(f"❌ Excel not found at {EXCEL_PATH}")
        return

    temp_path = EXCEL_PATH.replace(".xlsx", "_temp.xlsx")
    shutil.copyfile(EXCEL_PATH, temp_path)

    wb = openpyxl.load_workbook(temp_path)
    if EXCEL_SHEET_NAME not in wb.sheetnames:
        st.error("❌ Sheet name incorrect.")
        return

    ws = wb[EXCEL_SHEET_NAME]

    excel_headers = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            excel_headers[str(header_value).lower()] = col

    mapping = {
        "product description": fields.get("Product Description"),
        "lot #": fields.get("Batch/Lot No."),
        "date": fields.get("Date"),
        "sku": fields.get("SKU"),
        "qty": fields.get("Qty"),
    }

    target_row = EXCEL_INSERT_ROW
    while ws.max_row < target_row:
        ws.append([])

    for key, value in mapping.items():
        if key in excel_headers:
            col = excel_headers[key]
            ws.cell(row=target_row, column=col, value=value)

    wb.save(temp_path)
    shutil.move(temp_path, EXCEL_PATH)
    st.success("💾 Saved to Excel successfully!")
